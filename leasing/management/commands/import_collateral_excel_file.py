import argparse
import datetime
import decimal
import re
from decimal import Decimal

from dateutil.parser import parse, parserinfo
from django.core.management.base import BaseCommand

from leasing.models import Collateral, Contract, Lease

from openpyxl import load_workbook  # isort:skip (Not installed in CI or production)


LEASE_IDENTIFIER_COLUMN = 3
AMOUNT_COLUMN = 4
PAID_DATE_COLUMN = 5
RETURNED_DATE_COLUMN = 6
NOTE_COLUMN = 9


def get_date_from_value(date_value):
    if not date_value:
        return None

    if isinstance(date_value, datetime.datetime):
        return date_value.date()

    if isinstance(date_value, datetime.date):
        return date_value

    date = re.search(r'\d+\.\d+\.\d+', date_value, re.IGNORECASE)
    if date:
        return parse(date.group(0), parserinfo=parserinfo(dayfirst=True)).date()

    return None


class Command(BaseCommand):
    help = 'Import collaterals from excel (xlsx) file'

    def add_arguments(self, parser):
        parser.add_argument('excel_file', type=argparse.FileType('rb'),
                            help='Excel (xlsx) file with the collaterals')

    def handle(self, *args, **options):  # noqa
        from auditlog.registry import auditlog

        # Unregister models from auditlog
        auditlog.unregister(Collateral)

        wb = load_workbook(filename=options['excel_file'])
        sheet = wb.active

        max_row = sheet.max_row
        for row_num in range(2, max_row):
            lease_identifier_cell = sheet.cell(row=row_num, column=LEASE_IDENTIFIER_COLUMN)
            amount_cell = sheet.cell(row=row_num, column=AMOUNT_COLUMN)
            paid_date_cell = sheet.cell(row=row_num, column=PAID_DATE_COLUMN)
            returned_date_cell = sheet.cell(row=row_num, column=RETURNED_DATE_COLUMN)
            note_cell = sheet.cell(row=row_num, column=NOTE_COLUMN)

            if lease_identifier_cell.value is None:
                continue

            if lease_identifier_cell.value is None or amount_cell.value is None:
                amount = Decimal(0)
            else:
                try:
                    amount = Decimal(str(amount_cell.value))
                except decimal.InvalidOperation:
                    amount = Decimal(0)

            lease_identifier_cell_value = lease_identifier_cell.value.strip()
            # Fix typos in excel
            if lease_identifier_cell_value.startswith('S120-'):
                lease_identifier_cell_value = 'S0' + lease_identifier_cell_value[1:]

            if lease_identifier_cell_value[0:2] == 'So':
                lease_identifier_cell_value = 'S0' + lease_identifier_cell_value[2:]

            lease_identifiers = re.findall(r'[A-Z]\d{4}-\d+', lease_identifier_cell_value, re.IGNORECASE)
            if not lease_identifiers:
                continue

            paid_date = get_date_from_value(paid_date_cell.value)
            if not paid_date:
                # Try to find date from the lease identifier column
                paid_date = get_date_from_value(lease_identifier_cell.value)

            returned_date = get_date_from_value(returned_date_cell.value)

            for lease_identifier in lease_identifiers:
                try:
                    lease = Lease.objects.get_by_identifier(lease_identifier.upper())
                except Lease.DoesNotExist:
                    self.stderr.write('Lease "{}" not found'.format(lease_identifier))
                    continue

                lease_contract = lease.contracts.filter(type=1).first()  # 1 = Vuokrasopimus
                if not lease_contract:
                    self.stdout.write('Lease "{}" no lease contract found. Creating.'.format(lease_identifier))
                    lease_contract = Contract.objects.create(
                        lease=lease,
                        type_id=1,
                        signing_note='Vakuuksien tuontia varten luotu tyhjä sopimus'
                    )

                (collateral, collateral_created) = Collateral.objects.get_or_create(
                    contract=lease_contract,
                    type_id=2,  # 2 = Rahavakuus
                    total_amount=amount,
                    paid_date=paid_date,
                    returned_date=returned_date,
                    note=note_cell.value
                )
